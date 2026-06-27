import openpyxl
from io import BytesIO
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q
from django.core.paginator import Paginator
from django.views.decorators.http import require_GET, require_POST
from django.contrib import messages
from .models import Ingfou

COLUMNS = [
    {'field': 'sous_secteur', 'label': 'SOUS_SECTEUR', 'type': 'text'},
    {'field': 'n_operati', 'label': 'N_OPERATI', 'type': 'number'},
    {'field': 'chapitre', 'label': 'CHAPITRE_', 'type': 'number'},
    {'field': 'libelle_op', 'label': 'LIBELLE_OP', 'type': 'text'},
    {'field': 'ap_initial', 'label': 'AP_INITIAL', 'type': 'decimal'},
    {'field': 'commune', 'label': 'COMMUNE', 'type': 'text'},
    {'field': 'gest', 'label': 'GEST_', 'type': 'text'},
]


def _build_filter(q_dict):
    conditions = Q()
    for key, val in q_dict.items():
        if val == '':
            continue
        f = Ingfou._meta.get_field(key)
        if f is None:
            continue
        if f.get_internal_type() in ('IntegerField', 'DecimalField'):
            try:
                v = int(val) if f.get_internal_type() == 'IntegerField' else float(val)
                conditions &= Q(**{key: v})
            except ValueError:
                conditions &= Q(**{f'{key}__icontains': val})
        else:
            conditions &= Q(**{f'{key}__icontains': val})
    return conditions


def _advanced_build_filter(q_dict):
    conditions = Q()
    for key, val in q_dict.items():
        if val == '':
            continue
        field = key

        if val.startswith('!='):
            v = val[2:].strip()
            conditions &= ~Q(**{field: v}) if v else conditions
        elif val.startswith('>='):
            v = val[2:].strip()
            try:
                conditions &= Q(**{f'{field}__gte': int(v)})
            except ValueError:
                conditions &= Q(**{f'{field}__gte': v})
        elif val.startswith('<='):
            v = val[2:].strip()
            try:
                conditions &= Q(**{f'{field}__lte': int(v)})
            except ValueError:
                conditions &= Q(**{f'{field}__lte': v})
        elif val.startswith('>'):
            v = val[1:].strip()
            try:
                conditions &= Q(**{f'{field}__gt': int(v)})
            except ValueError:
                conditions &= Q(**{f'{field}__gt': v})
        elif val.startswith('<'):
            v = val[1:].strip()
            try:
                conditions &= Q(**{f'{field}__lt': int(v)})
            except ValueError:
                conditions &= Q(**{f'{field}__lt': v})
        elif val.startswith('='):
            v = val[1:].strip()
            conditions &= Q(**{field: v})
        else:
            conditions &= Q(**{f'{field}__icontains': val})
    return conditions


def _get_filtered_queryset(request):
    q = request.GET.get('q', '')
    filters = {}
    for col in COLUMNS:
        v = request.GET.get(col['field'], '')
        if v:
            filters[col['field']] = v

    records = Ingfou.objects.all()
    if filters:
        records = records.filter(_advanced_build_filter(filters))
    elif q:
        records = records.filter(
            Q(sous_secteur__icontains=q) |
            Q(libelle_op__icontains=q) |
            Q(commune__icontains=q)
        )
    return records, filters, q


@require_GET
def index(request):
    return _render_index(request, 'treeviewDJ/index.html')


@require_GET
def index1(request):
    return _render_index(request, 'treeviewDJ/index1.html')


@require_GET
def index2(request):
    return _render_index(request, 'treeviewDJ/index2.html')


@require_GET
def index3(request):
    return _render_index(request, 'treeviewDJ/index3.html')


def _render_index(request, template):
    records, filters, q = _get_filtered_queryset(request)

    sort = request.GET.get('sort', 'id')
    direction = request.GET.get('dir', 'asc')
    if sort not in [c['field'] for c in COLUMNS] + ['id']:
        sort = 'id'
    if direction == 'desc':
        records = records.order_by(f'-{sort}')
    else:
        records = records.order_by(sort)

    per_page = request.GET.get('per_page', 50)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 50
    per_page = max(10, min(200, per_page))

    paginator = Paginator(records, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    total = paginator.count

    return render(request, template, {
        'page_obj': page_obj,
        'columns': COLUMNS,
        'total': total,
        'q': q,
        'filters': filters,
        'sort': sort,
        'direction': direction,
        'per_page': per_page,
    })


@require_POST
def add_record(request):
    data = request.POST
    Ingfou.objects.create(
        sous_secteur=data.get('sous_secteur') or None,
        n_operati=data.get('n_operati') or None,
        chapitre=data.get('chapitre') or None,
        libelle_op=data.get('libelle_op') or None,
        ap_initial=data.get('ap_initial') or None,
        commune=data.get('commune') or None,
        gest=data.get('gest') or None,
    )
    messages.success(request, 'Enregistrement ajouté avec succès.')
    return redirect(request.POST.get('next', 'index'))


@require_GET
def get_record(request, pk):
    obj = get_object_or_404(Ingfou, pk=pk)
    return JsonResponse({
        'id': obj.id,
        'sous_secteur': obj.sous_secteur or '',
        'n_operati': obj.n_operati or '',
        'chapitre': obj.chapitre or '',
        'libelle_op': obj.libelle_op or '',
        'ap_initial': str(obj.ap_initial) if obj.ap_initial is not None else '',
        'commune': obj.commune or '',
        'gest': obj.gest or '',
    })


@require_POST
def edit_record(request, pk):
    obj = get_object_or_404(Ingfou, pk=pk)
    data = request.POST
    obj.sous_secteur = data.get('sous_secteur') or None
    obj.n_operati = data.get('n_operati') or None
    obj.chapitre = data.get('chapitre') or None
    obj.libelle_op = data.get('libelle_op') or None
    obj.ap_initial = data.get('ap_initial') or None
    obj.commune = data.get('commune') or None
    obj.gest = data.get('gest') or None
    obj.save()
    messages.success(request, 'Enregistrement modifié avec succès.')
    return redirect(request.POST.get('next', 'index'))


@require_POST
def delete_records(request):
    ids = request.POST.getlist('ids')
    if ids:
        pk_list = [int(x) for x in ids[0].split(',') if x.strip().isdigit()]
        count = Ingfou.objects.filter(pk__in=pk_list).delete()[0]
        messages.success(request, f'{count} enregistrement(s) supprimé(s).')
    return redirect(request.POST.get('next', 'index'))


@require_POST
def delete_all(request):
    count = Ingfou.objects.all().delete()[0]
    messages.success(request, f'Totalité des {count} enregistrements supprimée.')
    return redirect('index')


@require_GET
def export_excel(request):
    records, filters, q = _get_filtered_queryset(request)

    sort = request.GET.get('sort', 'id')
    direction = request.GET.get('dir', 'asc')
    if sort not in [c['field'] for c in COLUMNS] + ['id']:
        sort = 'id'
    records = records.order_by(f'-{sort}' if direction == 'desc' else sort)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ingfou"

    headers = [col['label'] for col in COLUMNS]
    ws.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        cell.fill = openpyxl.styles.PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')

    for r in records:
        ws.append([
            r.sous_secteur or '',
            r.n_operati or '',
            r.chapitre or '',
            r.libelle_op or '',
            float(r.ap_initial) if r.ap_initial is not None else '',
            r.commune or '',
            r.gest or '',
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 30)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ingfou_table.xlsx"'
    wb.save(response)
    return response


@require_GET
def count_records(request):
    records, filters, q = _get_filtered_queryset(request)
    total = records.count()
    return JsonResponse({'total': total})


@require_POST
def import_excel(request):
    file = request.FILES.get('file')
    if not file:
        messages.error(request, 'Veuillez sélectionner un fichier.')
        return redirect('index')

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            messages.error(request, 'Le fichier est vide.')
            return redirect('index')

        header_row = [str(c).strip().lower() if c else '' for c in rows[0]]
        col_map = {}
        for col in COLUMNS:
            for i, h in enumerate(header_row):
                if h == col['label'].lower() or h == col['field'].lower():
                    col_map[col['field']] = i
                    break

        if not col_map:
            messages.error(request, 'Aucune colonne reconnue. Vérifiez les en-têtes.')
            return redirect('index')

        created = 0
        for row in rows[1:]:
            if not any(cell is not None for cell in row):
                continue
            data = {}
            for field, idx in col_map.items():
                val = row[idx] if idx < len(row) else None
                if val is None or val == '':
                    data[field] = None
                elif field in ('n_operati', 'chapitre'):
                    try:
                        data[field] = int(float(str(val)))
                    except (ValueError, TypeError):
                        data[field] = None
                elif field == 'ap_initial':
                    try:
                        data[field] = float(str(val))
                    except (ValueError, TypeError):
                        data[field] = None
                else:
                    data[field] = str(val)
            Ingfou.objects.create(**data)
            created += 1

        messages.success(request, f'{created} enregistrements importés depuis Excel.')
    except Exception as e:
        messages.error(request, f'Erreur lors de l\'import : {e}')

    return redirect('index')
